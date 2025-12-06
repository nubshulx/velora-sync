"""
Excel document handler for test case management
"""

from pathlib import Path
from typing import Optional, List, Dict, Any
import openpyxl
from openpyxl.workbook import Workbook
from openpyxl.worksheet.worksheet import Worksheet
from openpyxl.styles import Alignment, Font
from datetime import datetime
import tempfile
import shutil

from src.utils.exceptions import DocumentReadError, DocumentWriteError
from src.utils.logger import get_logger
from src.document_readers.sharepoint_client import SharePointClient

logger = get_logger(__name__)


class ExcelHandler:
    """Handler for Excel test case documents"""
    
    def __init__(
        self,
        template: Dict[str, str],
        sharepoint_client: Optional[SharePointClient] = None,
        create_backup: bool = True
    ):
        """
        Initialize Excel handler
        
        Args:
            template: Test case template dictionary
            sharepoint_client: Optional SharePoint client
            create_backup: Whether to create backup before updates
        """
        self.template = template
        self.sharepoint_client = sharepoint_client
        self.create_backup = create_backup
        
        # Add timestamp columns to template
        self.columns = list(template.keys()) + ['Created', 'Updated']
    
    def read_test_cases(self, document_path: str) -> List[Dict[str, Any]]:
        """
        Read existing test cases from Excel
        
        Args:
            document_path: Path to Excel document
            
        Returns:
            List of test case dictionaries
            
        Raises:
            DocumentReadError: If reading fails
        """
        try:
            # Get local file path
            local_path = self._get_local_path(document_path, download=True)
            
            if not local_path.exists():
                logger.info("Excel file does not exist, will create new one")
                return []
            
            logger.info(f"Reading test cases from Excel: {document_path}")
            
            workbook = openpyxl.load_workbook(local_path)
            sheet = workbook.active
            
            # Read header row
            headers = [cell.value for cell in sheet[1]]
            
            # Read test cases
            test_cases = []
            for row in sheet.iter_rows(min_row=2, values_only=True):
                if not any(row):  # Skip empty rows
                    continue
                
                test_case = {}
                for idx, header in enumerate(headers):
                    if idx < len(row) and header:
                        test_case[header] = row[idx]
                
                test_cases.append(test_case)
            
            workbook.close()
            logger.info(f"Read {len(test_cases)} test cases from Excel")
            return test_cases
            
        except Exception as e:
            logger.error(f"Failed to read Excel document: {str(e)}")
            raise DocumentReadError(f"Failed to read Excel: {str(e)}")
    
    def write_test_cases(
        self,
        document_path: str,
        test_cases: List[Dict[str, Any]],
        mode: str = 'new_only'
    ) -> Dict[str, int]:
        """
        Write test cases to Excel
        
        Args:
            document_path: Path to Excel document
            test_cases: List of test case dictionaries
            mode: Update mode ('new_only' or 'full_sync')
            
        Returns:
            Dictionary with statistics (created, updated, total)
            
        Raises:
            DocumentWriteError: If writing fails
        """
        try:
            logger.info(f"Writing test cases to Excel in {mode} mode")
            
            # Get local file path
            local_path = self._get_local_path(document_path, download=True)
            
            # Create backup if file exists
            if self.create_backup and local_path.exists():
                backup_path = local_path.with_suffix('.xlsx.backup')
                shutil.copy2(local_path, backup_path)
                logger.info(f"Created backup: {backup_path}")
            
            # Read existing test cases
            existing_test_cases = self.read_test_cases(document_path) if local_path.exists() else []
            
            # Merge test cases based on mode
            merged_cases, stats = self._merge_test_cases(
                existing_test_cases,
                test_cases,
                mode
            )
            
            # Create new workbook
            workbook = Workbook()
            sheet = workbook.active
            sheet.title = "Test Cases"
            
            # Define fonts
            header_font = Font(name='Aptos Display', size=11, bold=True)
            body_font = Font(name='Aptos Display', size=11)
            
            # Write headers with bold font
            for col_idx, column in enumerate(self.columns, start=1):
                cell = sheet.cell(row=1, column=col_idx, value=column)
                cell.font = header_font
                cell.alignment = Alignment(horizontal='center', vertical='center')
            
            # Write test cases with Aptos Display font
            for row_idx, test_case in enumerate(merged_cases, start=2):
                for col_idx, column in enumerate(self.columns, start=1):
                    value = test_case.get(column, '')
                    cell = sheet.cell(row=row_idx, column=col_idx, value=value)
                    cell.font = body_font
                    # Enable text wrapping for cells with multi-line content
                    if value and '\n' in str(value):
                        cell.alignment = Alignment(wrap_text=True, vertical='top')
                    else:
                        cell.alignment = Alignment(vertical='top')
            
            # Auto-adjust column widths
            for column in sheet.columns:
                max_length = 0
                column_letter = column[0].column_letter
                for cell in column:
                    try:
                        if cell.value:
                            # Take max of first line for multi-line cells
                            first_line = str(cell.value).split('\n')[0]
                            max_length = max(max_length, len(first_line))
                    except:
                        pass
                adjusted_width = min(max_length + 2, 60)
                sheet.column_dimensions[column_letter].width = adjusted_width
            
            # Auto-adjust row heights for wrapped text
            for row in sheet.iter_rows(min_row=2, max_row=sheet.max_row):
                max_lines = 1
                for cell in row:
                    if cell.value and '\n' in str(cell.value):
                        lines = str(cell.value).count('\n') + 1
                        max_lines = max(max_lines, lines)
                # Set row height (approximately 15 points per line)
                sheet.row_dimensions[row[0].row].height = max_lines * 15
            
            # Save workbook
            workbook.save(local_path)
            workbook.close()
            
            logger.info(f"Saved {len(merged_cases)} test cases to Excel")
            
            # Upload to SharePoint if needed
            if 'sharepoint.com' in document_path.lower():
                if not self.sharepoint_client:
                    raise DocumentWriteError("SharePoint client not configured")
                
                logger.info("Uploading Excel to SharePoint")
                self.sharepoint_client.upload_file(local_path, document_path)
            
            return stats
            
        except DocumentWriteError:
            raise
        except Exception as e:
            logger.error(f"Failed to write Excel document: {str(e)}")
            raise DocumentWriteError(f"Failed to write Excel: {str(e)}")
    
    def _get_local_path(self, document_path: str, download: bool = False) -> Path:
        """
        Get local path for document (download from SharePoint if needed)
        
        Args:
            document_path: Document path or URL
            download: Whether to download from SharePoint
            
        Returns:
            Local file path
        """
        is_sharepoint = 'sharepoint.com' in document_path.lower()
        
        if is_sharepoint:
            if not self.sharepoint_client:
                raise DocumentReadError("SharePoint client not configured")
            
            # Use temporary file for SharePoint documents
            tmp_path = Path(tempfile.gettempdir()) / "velora_sync_excel.xlsx"
            
            if download:
                try:
                    self.sharepoint_client.download_file(document_path, tmp_path)
                except Exception as e:
                    # File might not exist yet, that's okay
                    logger.debug(f"Could not download Excel from SharePoint: {e}")
            
            return tmp_path
        else:
            return Path(document_path)
    
    def _merge_test_cases(
        self,
        existing: List[Dict[str, Any]],
        new: List[Dict[str, Any]],
        mode: str
    ) -> tuple[List[Dict[str, Any]], Dict[str, int]]:
        """
        Merge existing and new test cases based on mode
        
        Args:
            existing: Existing test cases
            new: New test cases
            mode: Update mode
            
        Returns:
            Tuple of (merged test cases, statistics)
        """
        stats = {'created': 0, 'updated': 0, 'unchanged': 0}
        current_time = datetime.now().strftime('%Y-%m-%d %H:%M:%S')
        
        if mode == 'new_only':
            # Only add new test cases, preserve existing ones
            merged = existing.copy()
            
            # Create index of existing test cases by ID
            existing_ids = {tc.get('Test Case ID', ''): tc for tc in existing}
            
            for test_case in new:
                tc_id = test_case.get('Test Case ID', '')
                if tc_id not in existing_ids:
                    # New test case
                    test_case['Created'] = current_time
                    test_case['Updated'] = current_time
                    merged.append(test_case)
                    stats['created'] += 1
                else:
                    stats['unchanged'] += 1
            
        elif mode == 'full_sync':
            # Update existing and add new test cases
            merged = []
            
            # Create index of existing test cases by ID
            existing_ids = {tc.get('Test Case ID', ''): tc for tc in existing}
            
            for test_case in new:
                tc_id = test_case.get('Test Case ID', '')
                
                if tc_id in existing_ids:
                    # Update existing test case
                    existing_tc = existing_ids[tc_id]
                    
                    # Check if content changed
                    content_changed = False
                    for key in test_case:
                        if key not in ['Created', 'Updated']:
                            if test_case.get(key) != existing_tc.get(key):
                                content_changed = True
                                break
                    
                    if content_changed:
                        # Preserve Created timestamp, update Updated
                        test_case['Created'] = existing_tc.get('Created', current_time)
                        test_case['Updated'] = current_time
                        merged.append(test_case)
                        stats['updated'] += 1
                    else:
                        # No changes, keep existing
                        merged.append(existing_tc)
                        stats['unchanged'] += 1
                    
                    # Remove from index
                    del existing_ids[tc_id]
                else:
                    # New test case
                    test_case['Created'] = current_time
                    test_case['Updated'] = current_time
                    merged.append(test_case)
                    stats['created'] += 1
            
            # Add remaining existing test cases that weren't in new list
            for existing_tc in existing_ids.values():
                merged.append(existing_tc)
                stats['unchanged'] += 1
        
        else:
            raise ValueError(f"Invalid mode: {mode}")
        
        stats['total'] = len(merged)
        return merged, stats
